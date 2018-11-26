#openpyxl笔记，用python操作excel
from openpyxl import Workbook
from openpyxl import load_workbook
# 创建一个工作簿
wb = Workbook()
ws = wb.active
wb.save('/usr/fishs/new.xlsx')
# 打开一个已经存在的工作簿
wb = load_workbook('/usr/fishc/old.xlsx')
ws = wb.active
ws1 = wb.get_sheet_by_name('sheet1')
# 对其中的某个单元格操作。例如：填写A3=hello
ws['A3'] = 'hello'
ws1['A3'] = 'hello'
# 获取其中某个单元格的值
A3 = ws['A3'].value
# 对表格循环操作
for lines in ws:
    for line in lines:
        print('the cell value:',line.value)
# 用数组的方式对表操作,行 Row 列 Column 。ws[row][col]
A1_value = ws[0][0]
# 总共有多上行,多少列
count_row = ws.max_row
count_column = ws.max_column